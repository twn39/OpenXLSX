import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "TestSheet"

# 1. Basic Values
ws["A1"] = "Hello from openpyxl"
ws["A2"] = 42
ws["A3"] = 3.1415

# 2. Date and Time (Crucial for serialization format checking)
ws["A4"] = datetime.datetime(2024, 5, 28, 14, 30, 0)

# 3. Formulas
ws["A5"] = "=SUM(A2:A3)"

# 4. Styling (Font, Fill, Border, Alignment)
ws["B1"] = "Styled"
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
ws["B1"].font = Font(bold=True, italic=True, color="FF0000", name="Comic Sans MS")
ws["B1"].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
ws["B1"].border = thin_border
ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

# 5. Merged Cells
ws.merge_cells('C1:E2')
ws['C1'] = 'Merged Area'

# 6. Multiple Sheets
ws2 = wb.create_sheet(title="SecondSheet")
ws2["A1"] = "Data in second sheet"

wb.save("../openpyxl_generated.xlsx")

